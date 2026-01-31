# src/data_manager.py
"""
数据管理模块：负责读取/写入 students.xlsx 与 credentials.json
使用 openpyxl 操作 xlsx 文件
"""
import os
import json
from openpyxl import Workbook, load_workbook

class DataManager:
    def __init__(self, data_dir):
        self.data_dir = data_dir
        self.students_file = os.path.join(self.data_dir, "students.xlsx")
        self.credentials_file = os.path.join(self.data_dir, "credentials.json")

    def ensure_initialized(self):
        # 创建 students.xlsx（若不存在）
        if not os.path.exists(self.students_file):
            self._create_sample_students()
        # 创建 credentials.json（若不存在），初始用户 zh / 0240
        if not os.path.exists(self.credentials_file):
            self._create_default_credentials()

    def _create_sample_students(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "students"
        headers = ["学号", "姓名", "性别", "班级"]
        ws.append(headers)
        sample = [
            ["2023001", "张三", "男", "一班"],
            ["2023002", "李四", "女", "一班"],
            ["2023003", "王五", "男", "二班"],
            ["2023004", "赵六", "女", "二班"],
        ]
        for row in sample:
            ws.append(row)
        wb.save(self.students_file)

    def _create_default_credentials(self):
        creds = {"username": "zh", "password": "0240"}
        with open(self.credentials_file, "w", encoding="utf-8") as f:
            json.dump(creds, f, ensure_ascii=False, indent=2)

    # 认证
    def validate_login(self, username, password):
        try:
            with open(self.credentials_file, "r", encoding="utf-8") as f:
                creds = json.load(f)
            return username == creds.get("username") and password == creds.get("password")
        except Exception:
            return False

    def change_password(self, new_username, new_password):
        creds = {"username": new_username, "password": new_password}
        with open(self.credentials_file, "w", encoding="utf-8") as f:
            json.dump(creds, f, ensure_ascii=False, indent=2)

    # 读取所有学生
    def load_all_students(self):
        wb = load_workbook(self.students_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        data = []
        for r in rows[1:]:
            record = dict(zip(headers, r))
            data.append(record)
        return data

    # 查找单个学生（按学号或姓名）
    def find_students(self, student_id=None, name=None, gender=None, class_name=None):
        all_students = self.load_all_students()
        results = []
        for s in all_students:
            if student_id and str(s.get("学号")) != str(student_id):
                continue
            if name and name not in str(s.get("姓名")):
                continue
            if gender and gender != "全部" and s.get("性别") != gender:
                continue
            if class_name and class_name != "全部" and s.get("班级") != class_name:
                continue
            results.append(s)
        return results

    # 统计按性别
    def stats_by_gender(self):
        all_students = self.load_all_students()
        total = len(all_students)
        counts = {}
        for s in all_students:
            g = s.get("性别", "未知")
            counts[g] = counts.get(g, 0) + 1
        # 计算比例
        stats = []
        for gender, cnt in counts.items():
            ratio = f"{cnt}/{total} ({cnt/total:.2%})" if total>0 else "0/0 (0.00%)"
            stats.append({"性别": gender, "人数": cnt, "比例": ratio})
        return stats

    # 导出统计到 xlsx（放到 data 文件夹）
    def export_stats(self, stats, filename="gender_stats.xlsx"):
        path = os.path.join(self.data_dir, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "gender_stats"
        ws.append(["性别", "人数", "比例"])
        for row in stats:
            ws.append([row["性别"], row["人数"], row["比例"]])
        wb.save(path)
        return path